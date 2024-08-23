import torch.nn as nn
import torch.nn.init as init
import os
from tqdm import tqdm
import torch
import numpy as np
import torch.utils.data as Data
from loss import *


def weights_init(m):
    if isinstance(m, nn.Linear):
        init.xavier_uniform_(m.weight)
        if m.bias is not None:
            init.zeros_(m.bias)


def modeltraining(model, LR, epoch_num, loss_mode, save_path, savestep, train_data, val_data, batchsize, PSF, reg,
                  SSIM_weight):
    if not os.path.exists(save_path):
        os.mkdir(save_path)

    train_loader = Data.DataLoader(
        dataset=train_data,
        batch_size=batchsize,
        shuffle=True,
        num_workers=1,
    )

    val_loader = Data.DataLoader(
        dataset=val_data,
        batch_size=1,
        shuffle=True,
        num_workers=1,
    )
    model.apply(weights_init)
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = model.to(device)
    criterionBCE = nn.BCELoss()
    PSF = PSF.to(device)
    optimizer = torch.optim.Adam(model.parameters(), lr=LR)
    criterion1 = nn.MSELoss()
    criterion2 = SSIM()
    criterion3 = PearsonCorrelation()

    if loss_mode == 1:
        BL = BlurryLoss()
    else:
        BL = BlurryLossMSE()
    MSE_vals = []
    SSIM_vals = []
    BL_vals = []
    BCE_vals = []
    MSE_trains = []
    SSIM_trains = []
    BCE_trains = []
    BL_trains = []
    best_val_loss = float('inf')

    for epoch in range(epoch_num):
        print(epoch)
        train_loss_epoch = 0
        val_loss_epoch = 0
        MSE_val_epoch = 0
        SSIM_val_epoch = 0
        MSE_train_epoch = 0
        SSIM_train_epoch = 0
        train_num = 0
        BCE_train_epoch = 0
        BL_train_epoch = 0
        BCE_val_epoch = 0
        BL_val_epoch = 0
        val_num = 0

        if epoch > epoch_num // 2 - 1:
            optimizer = torch.optim.Adam(model.parameters(), lr=LR / 10)
        train_loader = tqdm(train_loader, desc="Training epoch", total=len(train_loader))
        for step, (b_x, b_y) in enumerate(train_loader):
            b_x = b_x.to(device)
            b_y = b_y.to(device)
            b_y1, b_y2 = torch.chunk(b_y, 2, dim=1)
            model.train()
            output = model(b_x)
            # output = (output-output.min()) / (output.max()-output.min())
            MSEtrain = criterion1(output, b_y1)
            SSIMtrain = criterion2(output, b_y1)
            imgD = F.conv2d(output, PSF, padding=PSF.shape[2] // 2)
            PPCtrain = criterion3(imgD, b_y2)
            BLtrain = reg * (1 - PPCtrain)
            # BLtrain = reg * BL(output, b_y2, PSF)
            BCEtrain = reg * criterionBCE(output, b_y1)
            if loss_mode == 0:
                loss = MSEtrain + SSIM_weight * (1 - SSIMtrain)
            else:
                loss = MSEtrain + SSIM_weight * (1 - SSIMtrain) + BLtrain  # + BCEtrain

            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            train_loader.set_postfix_str(f"Epoch: {epoch + 1}, Loss: {loss.item()}")
            MSE_train_epoch += MSEtrain.detach() * b_x.size(0)
            SSIM_train_epoch += SSIMtrain.detach() * b_x.size(0)
            BL_train_epoch += BLtrain.detach() * b_x.size(0)
            BCE_train_epoch += BCEtrain.detach() * b_x.size(0)
            train_num = train_num + b_x.size(0)
            if step % savestep == 0:
                num_check = 0
                MSE_val_check = 0
                SSIM_val_check = 0
                BL_val_check = 0
                BCE_val_check = 0

                with torch.no_grad():
                    for step, (b_x, b_y) in enumerate(val_loader):
                        b_x = b_x.to(device)
                        b_y = b_y.to(device)
                        b_y1, b_y2 = torch.chunk(b_y, 2, dim=1)
                        model.eval()
                        output = model(b_x)
                        MSEval = criterion1(output, b_y1)
                        SSIMval = criterion2(output, b_y1)
                        BLval = reg * BL(output, b_y2, PSF)
                        BCEval = reg * criterionBCE(output, b_y1)
                        MSE_val_check += MSEval * b_x.size(0)
                        SSIM_val_check += SSIMval * b_x.size(0)
                        BL_val_check += BLval * b_x.size(0)
                        BCE_val_check += BCEval * b_x.size(0)
                        num_check = num_check + b_x.size(0)

                    MSE_val = MSE_val_check / num_check
                    SSIM_val = SSIM_val_check / num_check
                    BCE_val = BCE_val_check / num_check
                    BL_val = BL_val_check / num_check
                    val_loss = MSE_val + SSIM_weight * (1 - SSIM_val)

                    MSE_vals.append(MSE_val.to('cpu'))
                    SSIM_vals.append(SSIM_val.to('cpu'))
                    BCE_vals.append(BL_val.to('cpu'))
                    BL_vals.append(BL_val.to('cpu'))

                MSE_train = MSE_train_epoch / train_num
                SSIM_train = SSIM_train_epoch / train_num
                BCE_train = BCE_train_epoch / train_num
                BL_train = BL_train_epoch / train_num
                MSE_trains.append(MSE_train.to('cpu'))
                SSIM_trains.append(SSIM_train.to('cpu'))
                BCE_trains.append(BL_train.to('cpu'))
                BL_trains.append(BL_train.to('cpu'))
                if loss_mode == 0:
                    train_loss = MSE_train + SSIM_weight * (1 - SSIM_train)
                else:
                    train_loss = MSE_train + SSIM_weight * (1 - SSIM_train) + BL_train  # + BCE_train

        with torch.no_grad():
            for step, (b_x, b_y) in enumerate(val_loader):
                b_x = b_x.to(device)
                b_y = b_y.to(device)
                b_y1, b_y2 = torch.chunk(b_y, 2, dim=1)
                model.eval()
                output = model(b_x)
                MSEval = criterion1(output, b_y1)
                SSIMval = criterion2(output, b_y1)
                BLval = reg * BL(output, b_y2, PSF)
                BCEval = reg * criterionBCE(output, b_y1)

                MSE_val_epoch += MSEval * b_x.size(0)
                SSIM_val_epoch += SSIMval * b_x.size(0)
                BL_val_epoch += BLval * b_x.size(0)
                BCE_val_epoch += BCEval * b_x.size(0)
                val_num = val_num + b_x.size(0)

            MSE_val = MSE_val_epoch / val_num
            SSIM_val = SSIM_val_epoch / val_num
            BCE_val = BCE_val_epoch / train_num
            BL_val = BL_val_epoch / train_num
            # if loss_mode == 0:
            # val_loss = MSE_val + SSIM_weight * (1 - SSIM_val)
            # else:
            # val_loss = MSE_val + SSIM_weight * (1 - SSIM_val) + BCE_val + BL_val

            val_loss = MSE_val + SSIM_weight * (1 - SSIM_val)

        print(
            "Epoch: {}, Train Loss: {:.4f}, Train MSE: {:.4f}, Train SSIM: {:.4f}".format(epoch, train_loss, MSE_train,
                                                                                          SSIM_train))
        print("Epoch: {}, Val Loss: {:.4f}, Val MSE: {:.4f}, Val SSIM: {:.4f}".format(epoch, val_loss, MSE_val,
                                                                                      SSIM_val))

        torch.cuda.empty_cache()
        del train_loss_epoch, val_loss_epoch, MSE_val_epoch, SSIM_val_epoch, MSE_train_epoch, SSIM_train_epoch, train_num, val_num, b_x, b_y, b_y1, b_y2

    path = os.path.join(save_path, 'final_network.pth')
    torch.save(model.state_dict(), path)
    import openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MSE_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MSE_trains.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(SSIM_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'SSIM_trains.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(SSIM_vals)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'SSIM_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(MSE_vals)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'MSE_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(BCE_vals)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'BCE_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(BCE_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'BCE_trains.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(BL_vals)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'BL_vals.xlsx')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    a = np.array(BL_trains)
    for row in a:
        sheet.append([row])

    workbook.save(save_path + '\\' + 'BL_trains.xlsx')
    return model